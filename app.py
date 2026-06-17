"""
GDSS-Maverick Hackathon
AI-Driven Image-to-IMDB Tool
Streamlit UI
"""

import streamlit as st
from supabase import create_client
SUPABASE_URL = st.secrets.get("SUPABASE_URL", "")
SUPABASE_KEY = st.secrets.get("SUPABASE_KEY", "")
supabase_client = create_client(SUPABASE_URL, SUPABASE_KEY) if SUPABASE_URL and SUPABASE_KEY else None

import streamlit.components.v1 as components
import json

import re
import base64
import requests
import io
import os
import time
from pathlib import Path
from PIL import Image
from collections import defaultdict, Counter
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

# ─────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="IMDB Auto-Fill Tool",
    page_icon="🏷️",
    layout="wide"
)

# ─────────────────────────────────────────────
# STYLING
# ─────────────────────────────────────────────
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Playfair+Display:ital,wght@0,400..900;1,400..900&family=Outfit:wght@300;400;500;600;700;800&display=swap');

    .stApp, [data-testid="stAppViewContainer"], [data-testid="stApp"] {
        background-color: #f8fafc !important;
        color: #475569 !important;
        font-family: 'Outfit', sans-serif !important;
    }
    .main {
        background-color: transparent !important;
    }
    .block-container {
        padding-top: 6rem !important;
    }

    /* Typography */
    h1, h2, h3, h4, h5, h6, [data-testid="stHeader"] {
        font-family: 'Playfair Display', Georgia, serif !important;
        font-weight: 700 !important;
        color: #0f172a !important;
    }
    
    p, label, caption, small {
        font-family: 'Outfit', sans-serif !important;
        color: #475569 !important;
    }

    /* Sidebar elegant light theme */
    section[data-testid="stSidebar"] {
        background-color: #ffffff !important;
        border-right: 1px solid #e2e8f0 !important;
        box-shadow: 4px 0 20px rgba(0, 0, 0, 0.01) !important;
        z-index: 1000 !important; /* stay below our navbar (z-index: 99999) */
    }
    section[data-testid="stSidebar"] > div {
        background-color: #ffffff !important;
    }
    /* Push sidebar content below the floating navbar capsule */
    section[data-testid="stSidebar"] > div:first-child {
        padding-top: 72px !important;
    }
    section[data-testid="stSidebar"] h3 {
        font-size: 1.1rem !important;
        color: #0f172a !important;
        margin-top: 1rem !important;
    }
    /* ── Hide Streamlit's native sidebar toggle buttons ──
       Moved off-screen (NOT display:none or opacity:0) so they remain
       fully rendered and React's .click() always works. */
    [data-testid="stSidebarCollapsedControl"] {
        position: fixed !important;
        left: -9999px !important;
        top: -9999px !important;
    }
    section[data-testid="stSidebar"] [data-testid="stSidebarCollapseButton"] {
        position: fixed !important;
        left: -9999px !important;
        top: -9999px !important;
    }

    /* Input Fields & Widgets */
    div[data-testid="stTextInput"] [data-baseweb="input"] {
        background-color: #ffffff !important;
        border: 1px solid #cbd5e1 !important;
        border-radius: 10px !important;
        height: 42px !important;
        transition: border-color 0.15s ease, box-shadow 0.15s ease !important;
    }
    div[data-testid="stTextInput"] input {
        border: none !important;
        background-color: transparent !important;
        color: #0f172a !important;
        font-family: 'Outfit', sans-serif !important;
        height: 100% !important;
        box-shadow: none !important;
        outline: none !important;
    }
    div[data-testid="stSelectbox"] [data-baseweb="select"] {
        background-color: #ffffff !important;
        border: 1px solid #cbd5e1 !important;
        color: #0f172a !important;
        border-radius: 10px !important;
        font-family: 'Outfit', sans-serif !important;
        height: 42px !important;
        transition: border-color 0.15s ease, box-shadow 0.15s ease !important;
    }
    /* Target selectbox value wrapper to ensure it is white-bg and has dark text */
    div[data-testid="stSelectbox"] [data-baseweb="select"] > div {
        background-color: #ffffff !important;
        color: #0f172a !important;
    }
    /* Ensure child text and spans inside the selectbox also use the dark color */
    div[data-testid="stSelectbox"] [data-baseweb="select"] * {
        color: #0f172a !important;
    }
    div[data-testid="stTextInput"] [data-baseweb="input"]:hover,
    div[data-testid="stSelectbox"] [data-baseweb="select"]:hover {
        border-color: #94a3b8 !important;
    }
    div[data-testid="stTextInput"] [data-baseweb="input"]:focus-within,
    div[data-testid="stSelectbox"] [data-baseweb="select"]:focus {
        border-color: #0f172a !important;
        box-shadow: 0 0 0 1px #0f172a !important;
    }
    /* Placeholder colors */
    div[data-testid="stTextInput"] input::placeholder {
        color: #94a3b8 !important;
        opacity: 1 !important;
    }
    /* Selectbox dropdown items and portals */
    [data-baseweb="popover"] ul,
    [data-baseweb="menu"],
    div[role="listbox"] {
        background-color: #ffffff !important;
        color: #0f172a !important;
        border: 1px solid #e2e8f0 !important;
        border-radius: 10px !important;
    }
    [data-baseweb="popover"] ul li,
    [data-baseweb="menu"] li,
    div[role="option"],
    li[role="option"] {
        background-color: #ffffff !important;
        color: #0f172a !important;
        padding: 0.5rem 1rem !important;
        transition: background-color 0.15s ease !important;
    }
    [data-baseweb="popover"] ul li:hover,
    [data-baseweb="menu"] li:hover,
    div[role="option"]:hover,
    li[role="option"]:hover {
        background-color: #f1f5f9 !important;
        color: #0f172a !important;
    }
    div[data-testid="stSelectbox"] [data-baseweb="select"],
    div[role="option"],
    li[role="option"],
    [data-baseweb="menu"] * {
        cursor: pointer !important;
    }
    div[data-testid="stTextInput"] label, div[data-testid="stSelectbox"] label {
        font-family: 'Outfit', sans-serif !important;
        font-weight: 600 !important;
        color: #475569 !important;
        font-size: 0.85rem !important;
    }

    .nav-wrapper {
        container-type: inline-size;
        width: 100%;
    }

    /* Top Navigation Capsule */
    .nav-container {
        display: flex;
        align-items: center;
        justify-content: space-between;
        background: rgba(255, 255, 255, 0.55) !important;
        -webkit-backdrop-filter: blur(40px) saturate(200%) !important;
        backdrop-filter: blur(40px) saturate(200%) !important;
        border: 1px solid rgba(255, 255, 255, 0.75) !important;
        border-radius: 24px !important;
        padding: 0.6rem 1.75rem !important;
        box-shadow: 
            0 20px 40px rgba(15, 23, 42, 0.06), 
            0 5px 15px rgba(15, 23, 42, 0.04), 
            inset 0 1px 0 rgba(255, 255, 255, 0.7) !important;
        position: fixed !important;
        top: 12px !important;
        left: 50% !important;
        transform: translateX(-50%) !important;
        width: 92% !important;
        max-width: 1100px !important;
        z-index: 99999 !important;
        margin: 0 !important;
    }
    /* Mobile/Small Container: compact single-row navbar */
    @container (max-width: 850px) {
        .nav-container {
            width: 96% !important;
            padding: 0.4rem 0.75rem !important;
            border-radius: 16px !important;
        }
        .nav-logo {
            font-size: 0.85rem !important;
            margin-right: 0.5rem !important;
        }
        .nav-hamburger {
            width: 32px !important;
            height: 32px !important;
            margin-right: 0.5rem !important;
            border-radius: 8px !important;
        }
        /* Hide step text labels on mobile/small containers, show only the circled numbers */
        .nav-tab {
            font-size: 0 !important;
        }
        .step-num {
            font-size: 0.65rem !important;
            margin-right: 0 !important;
        }
        .nav-center {
            gap: 0.5rem !important;
        }
        .nav-badge {
            font-size: 0.65rem !important;
            padding: 0.2rem 0.5rem !important;
        }
    }

    /* Tablet & Medium Containers: slightly more compact navbar to prevent overflow */
    @container (min-width: 851px) and (max-width: 1100px) {
        .nav-container {
            padding: 0.5rem 1rem !important;
        }
        .nav-center {
            gap: 0.8rem !important;
        }
        .nav-tab {
            font-size: 0.82rem !important;
        }
        .step-num {
            width: 18px !important;
            height: 18px !important;
            font-size: 0.68rem !important;
            margin-right: 5px !important;
        }
        .nav-logo {
            font-size: 1.1rem !important;
        }
        .nav-badge {
            font-size: 0.75rem !important;
            padding: 0.3rem 0.6rem !important;
        }
    }

    /* Maintain viewport-based padding for the main content block container */
    @media (max-width: 950px) {
        .block-container {
            padding-top: 4.5rem !important;
        }
    }
    /* Hide the zero-height iframe container used for JS injection so it doesn't add blank space */
    div.element-container:has(iframe[height="0"]) {
        display: none !important;
        margin: 0 !important;
        padding: 0 !important;
    }
    
    /* Hide Streamlit's ⋮ mobile menu button */
    [data-testid="stMainMenuButton"],
    button[data-testid="baseButton-header"] {
        display: none !important;
    }
    /* Hamburger toggle button inside the navbar capsule */
    .nav-hamburger {
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        width: 40px !important;
        height: 40px !important;
        background: #000000 !important;
        border-radius: 12px !important;
        cursor: pointer !important;
        margin-right: 1.25rem !important;
        flex-shrink: 0 !important;
        transition: background 0.2s ease, transform 0.15s ease !important;
        user-select: none !important;
    }
    .nav-hamburger:hover {
        background: #000000 !important;
        transform: scale(1.05) !important;
    }
    /* Hide Streamlit's native header bar (Deploy btn, ⋮ menu) */
    header[data-testid="stHeader"] {
        display: none !important;
    }
    .nav-left {
        display: flex;
        align-items: center;
    }
    .nav-logo {

        font-family: 'Playfair Display', Georgia, serif !important;
        font-weight: 800 !important;
        font-size: 1.25rem !important;
        color: #0f172a !important;
    }
    .nav-center {
        display: flex;
        align-items: center;
        gap: 1.5rem;
    }
    .nav-tab {
        font-family: 'Outfit', sans-serif !important;
        font-size: 0.9rem !important;
        font-weight: 500 !important;
        color: #000000 !important;
        opacity: 0.55 !important;
        padding: 0.5rem 0.25rem !important;
        cursor: pointer;
        position: relative;
        transition: all 0.2s ease !important;
        display: inline-flex !important;
        align-items: center !important;
    }
    .nav-tab:hover {
        opacity: 0.95 !important;
    }
    .active-tab {
        color: #000000 !important;
        font-weight: 700 !important;
        opacity: 1 !important;
    }
    .active-tab::after {
        content: '';
        position: absolute;
        bottom: 0;
        left: 0;
        right: 0;
        height: 2.5px;
        background-color: #000000;
        border-radius: 2px;
    }
    .step-num {
        display: inline-flex !important;
        align-items: center !important;
        justify-content: center !important;
        width: 20px !important;
        height: 20px !important;
        background-color: rgba(0, 0, 0, 0.08) !important;
        color: #000000 !important;
        border-radius: 50% !important;
        font-size: 0.72rem !important;
        font-weight: 600 !important;
        margin-right: 8px !important;
        transition: all 0.2s ease !important;
        opacity: 0.8 !important;
    }
    .active-num {
        background-color: #000000 !important;
        color: #ffffff !important;
        opacity: 1 !important;
        box-shadow: 0 2px 6px rgba(0, 0, 0, 0.2) !important;
    }
    .nav-right {
        display: flex;
        align-items: center;
        gap: 1rem;
    }
    .nav-badge {
        font-family: 'Outfit', sans-serif !important;
        font-size: 0.8rem !important;
        font-weight: 600 !important;
        color: #475569 !important;
        background-color: #f1f5f9 !important;
        padding: 0.35rem 0.75rem !important;
        border-radius: 20px !important;
        display: flex;
        align-items: center;
        gap: 0.25rem;
        border: 1px solid #e2e8f0 !important;
    }
    .nav-profile {
        display: flex;
        align-items: center;
        gap: 0.5rem;
        background: #f8fafc !important;
        border: 1px solid #e2e8f0 !important;
        padding: 0.35rem 0.85rem !important;
        border-radius: 20px !important;
    }
    .nav-profile-name {
        font-family: 'Outfit', sans-serif !important;
        font-size: 0.85rem !important;
        font-weight: 600 !important;
        color: #0f172a !important;
    }

    /* Cards & Containers */
    div[data-testid="stMetric"], 
    div[data-testid="stAlert"] {
        background-color: #ffffff !important;
        border: 1px solid #e2e8f0 !important;
        border-radius: 20px !important;
        box-shadow: 0 10px 30px rgba(0, 0, 0, 0.02) !important;
        color: #475569 !important;
        transition: all 0.3s ease !important;
    }
    div[data-testid="stMetric"]:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 15px 35px rgba(0, 0, 0, 0.04) !important;
    }
    div[data-testid="stMetric"] {
        padding: 1.25rem !important;
    }
    div[data-testid="stMetricLabel"] > div {
        font-family: 'Outfit', sans-serif !important;
        text-transform: uppercase !important;
        font-size: 0.75rem !important;
        font-weight: 600 !important;
        letter-spacing: 0.05em !important;
        color: #64748b !important;
    }
    div[data-testid="stMetricValue"] > div {
        font-family: 'Playfair Display', Georgia, serif !important;
        font-size: 2.25rem !important;
        font-weight: 800 !important;
        color: #0f172a !important;
        margin-top: 0.25rem !important;
    }

    /* Expander & Bordered Containers */
    div[data-testid="stExpander"],
    div[data-testid="stVerticalBlockBorderWrapper"] {
        background-color: #ffffff !important;
        border: 1px solid #e2e8f0 !important;
        border-radius: 16px !important;
        box-shadow: 0 10px 30px rgba(0, 0, 0, 0.02) !important;
        color: #475569 !important;
        transition: all 0.3s ease !important;
        overflow: hidden !important;
        margin-bottom: 1.5rem !important;
    }
    div[data-testid="stExpander"]:hover,
    div[data-testid="stVerticalBlockBorderWrapper"]:hover {
        border-color: #cbd5e1 !important;
        box-shadow: 0 12px 35px rgba(0, 0, 0, 0.03) !important;
    }

    /* Expander Header (Modern Streamlit uses summary element) */
    div[data-testid="stExpander"] summary,
    .streamlit-expanderHeader {
        background-color: #ffffff !important;
        border-bottom: none !important;
        font-family: 'Outfit', sans-serif !important;
        font-weight: 600 !important;
        color: #0f172a !important;
        padding: 1rem 1.5rem !important;
        transition: background-color 0.2s ease !important;
    }
    /* Expander Header Hover */
    div[data-testid="stExpander"] summary:hover {
        background-color: #f8fafc !important;
    }
    /* Target the text wrapper inside summary */
    div[data-testid="stExpander"] summary > div[data-testid="stMarkdownContainer"] p {
        color: #0f172a !important;
        font-family: 'Outfit', sans-serif !important;
        font-weight: 600 !important;
    }
    /* Target the toggle icon */
    div[data-testid="stExpander"] summary > svg[data-testid="stExpanderToggleIcon"] {
        fill: #0f172a !important;
    }

    /* Expander Content */
    div[data-testid="stExpander"] div[data-testid="stExpanderDetails"],
    .streamlit-expanderContent {
        background-color: #ffffff !important;
        border-top: 1px solid #f1f5f9 !important;
        padding: 1.5rem !important;
    }

    /* Buttons */
    div.stButton > button, div.stDownloadButton > button {
        background-color: #000000 !important;
        color: #ffffff !important;
        border: none !important;
        border-radius: 12px !important;
        padding: 0.65rem 1.75rem !important;
        font-family: 'Outfit', sans-serif !important;
        font-weight: 600 !important;
        font-size: 0.95rem !important;
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.15) !important;
        transition: all 0.2s ease !important;
        width: 100% !important;
    }
    /* Force white on all nested label elements Streamlit injects inside buttons */
    div.stButton > button *, div.stDownloadButton > button * {
        color: #ffffff !important;
    }
    div.stButton > button:hover, div.stDownloadButton > button:hover {
        background-color: #222222 !important;
        box-shadow: 0 6px 16px rgba(0, 0, 0, 0.25) !important;
        transform: translateY(-1px) !important;
    }
    div.stButton > button:active, div.stDownloadButton > button:active {
        transform: translateY(0px) !important;
    }


    /* File Uploader dashed zone */
    div[data-testid="stFileUploader"] section {
        border: 2px dashed #cbd5e1 !important;
        background-color: #ffffff !important;
        border-radius: 20px !important;
        padding: 2rem !important;
        transition: all 0.2s ease !important;
        box-shadow: 0 4px 20px rgba(0, 0, 0, 0.005) !important;
    }
    div[data-testid="stFileUploader"] section:hover {
        border-color: #0f172a !important;
        background-color: #f8fafc !important;
    }
    /* Fix file uploader text elements that default to white/invisible in dark theme base */
    div[data-testid="stFileUploader"] section span,
    div[data-testid="stFileUploader"] section p,
    div[data-testid="stFileUploader"] section small,
    div[data-testid="stFileUploader"] section div {
        color: #475569 !important;
    }
    /* Ensure the Upload button inside file uploader keeps its dark background and white text */
    div[data-testid="stFileUploader"] section button,
    div[data-testid="stFileUploader"] section button * {
        color: #ffffff !important;
        background-color: #000000 !important;
    }
    div[data-testid="stFileUploader"] section button:hover {
        background-color: #222222 !important;
    }

    /* Table & Dataframe */
    div[data-testid="stDataFrame"] {
        border: 1px solid #e2e8f0 !important;
        border-radius: 16px !important;
        overflow: hidden !important;
        box-shadow: 0 4px 20px rgba(0, 0, 0, 0.01) !important;
    }

    /* Success banner */
    .success-banner {
        background-color: #ecfdf5 !important;
        border: 1px solid #a7f3d0 !important;
        border-radius: 12px !important;
        color: #065f46 !important;
        font-family: 'Outfit', sans-serif !important;
        font-weight: 600 !important;
        padding: 1rem !important;
        text-align: center !important;
        margin-top: 1rem;
        margin-bottom: 1rem;
    }

    /* Camera Input Widget styling */
    div[data-testid="stCameraInput"] {
        border: 1px solid #cbd5e1 !important;
        border-radius: 20px !important;
        overflow: hidden !important;
        box-shadow: 0 10px 30px rgba(0, 0, 0, 0.02) !important;
        background-color: #ffffff !important;
        margin-bottom: 1rem !important;
    }
    div[data-testid="stCameraInput"] button,
    div[data-testid="stCameraInput"] button * {
        color: #ffffff !important;
        background-color: #000000 !important;
        border-radius: 10px !important;
    }
    div[data-testid="stCameraInput"] button:hover {
        background-color: #222222 !important;
    }

    /* Compact Snapshot Gallery Grid */
    .snap-card-title {
        font-family: 'Outfit', sans-serif !important;
        font-size: 0.8rem !important;
        font-weight: 600 !important;
        color: #475569 !important;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
        margin-top: 0.25rem !important;
        margin-bottom: 0.5rem !important;
    }

    /* Navbar clickable view links */
    .nav-view-link {
        font-family: 'Outfit', sans-serif !important;
        font-size: 0.88rem !important;
        font-weight: 600 !important;
        color: #000000 !important;
        opacity: 0.5 !important;
        padding: 0.45rem 0.85rem !important;
        border-radius: 10px !important;
        cursor: pointer !important;
        text-decoration: none !important;
        transition: all 0.2s ease !important;
        display: inline-flex !important;
        align-items: center !important;
        gap: 0.35rem !important;
    }
    .nav-view-link:hover {
        opacity: 0.85 !important;
        background-color: rgba(0,0,0,0.05) !important;
    }
    .nav-view-link.active-view {
        opacity: 1 !important;
        background-color: #000000 !important;
        color: #ffffff !important;
    }
    .nav-divider {
        width: 1px;
        height: 20px;
        background: #e2e8f0;
        margin: 0 0.5rem;
    }
    /* Item Master Database table card */
    .db-card {
        background: #ffffff;
        border: 1px solid #e2e8f0;
        border-radius: 16px;
        padding: 1.5rem;
        box-shadow: 0 4px 20px rgba(0,0,0,0.02);
        margin-bottom: 1.5rem;
    }
    .db-stat-chip {
        display: inline-flex;
        align-items: center;
        gap: 0.4rem;
        background: #f1f5f9;
        border: 1px solid #e2e8f0;
        border-radius: 20px;
        padding: 0.3rem 0.75rem;
        font-family: 'Outfit', sans-serif;
        font-size: 0.8rem;
        font-weight: 600;
        color: #475569;
        margin-right: 0.5rem;
        margin-bottom: 0.5rem;
    }

    /* Stepper card in Pipeline Page */
    .stepper-card {
        background: #ffffff;
        border: 1px solid #e2e8f0;
        border-radius: 20px;
        padding: 0.85rem 1.5rem;
        box-shadow: 0 4px 20px rgba(15, 23, 42, 0.015);
        margin-bottom: 2rem;
        width: 100%;
    }
    .stepper-container {
        display: flex;
        align-items: center;
        justify-content: center;
        gap: 1.25rem;
        flex-wrap: wrap;
        width: 100%;
    }
    .page-tab {
        font-family: 'Outfit', sans-serif !important;
        font-size: 0.95rem !important;
        font-weight: 500 !important;
        color: #475569 !important;
        opacity: 0.55 !important;
        padding: 0.4rem 0.8rem !important;
        display: inline-flex !important;
        align-items: center !important;
        border-radius: 10px;
        transition: all 0.2s ease !important;
    }
    .page-tab.active-tab {
        color: #0f172a !important;
        font-weight: 700 !important;
        opacity: 1 !important;
        background-color: #f1f5f9;
        border: 1px solid #cbd5e1;
    }
    .stepper-arrow {
        color: #cbd5e1;
        font-size: 1rem;
        font-weight: 500;
        user-select: none;
    }

</style>
""", unsafe_allow_html=True)


def render_navbar(active_step, model_name="Groq API", current_view="pipeline"):
    pipeline_active = 'active-view' if current_view == 'pipeline' else ''
    database_active = 'active-view' if current_view == 'database' else ''

    navbar_html = f"""
    <div class="nav-wrapper">
        <div class="nav-container">
            <div class="nav-left">
                <span class="nav-logo">GDSS</span>
                <div class="nav-divider" style="margin-left:1rem;"></div>
                <a href="?view=pipeline" target="_self" class="nav-view-link {pipeline_active}" style="margin-left:0.5rem;">🏭 Pipeline</a>
                <a href="?view=database" target="_self" class="nav-view-link {database_active}">📊 Item Master Database</a>
            </div>
            <div class="nav-right">
                <span class="nav-badge"><span style="color: #22c55e; margin-right: 6px; font-size: 1rem; line-height: 1;">●</span>{model_name}</span>
            </div>
        </div>
    </div>
    """
    st.markdown(navbar_html, unsafe_allow_html=True)


def render_pipeline_stepper(active_step):
    steps = [
        ("Upload", 1),
        ("Extract", 2),
        ("Aggregate", 3),
        ("Validate", 4),
        ("Export", 5)
    ]
    tabs_html = ""
    for i, (label, step_num) in enumerate(steps):
        if step_num == active_step:
            tabs_html += f'<div class="page-tab active-tab"><span class="step-num active-num">{step_num}</span>{label}</div>'
        else:
            tabs_html += f'<div class="page-tab"><span class="step-num">{step_num}</span>{label}</div>'
        if i < len(steps) - 1:
            tabs_html += '<span class="stepper-arrow">→</span>'

    stepper_html = f"""
    <div class="stepper-card">
        <div class="stepper-container">
            {tabs_html}
        </div>
    </div>
    """
    st.markdown(stepper_html, unsafe_allow_html=True)



# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────
IMDB_COLS = [
    "ITEM NAME", "BARCODE", "MANUFACTURER", "BRAND", "WEIGHT",
    "PACKAGING TYPE", "COUNTRY", "VARIANT", "TYPE",
    "FRAGRANCE FLAVOR", "PROMOTION", "ADDONS", "TAGLINE",
]


SYSTEM_PROMPT = """You are a product data extraction engine for a retail catalog system.
YOUR ONLY OUTPUT IS A SINGLE VALID JSON OBJECT. NO PREAMBLE. NO EXPLANATION. NO MARKDOWN. NO CODE BLOCKS. JUST THE RAW JSON.

Extract the following 13 fields:
- ITEM_NAME: full descriptive product name for catalog e.g. "KNORR CHICKEN STOCK CUBE 20G"
- BARCODE: numeric digits only, no dashes or spaces
- MANUFACTURER: full company name that manufactures the product
- BRAND: brand name exactly as shown on package e.g. "KNORR", "SUNLIGHT"
- WEIGHT: net weight or volume with unit, no spaces e.g. "80G", "500ML", "1.5KG"
- PACKAGING_TYPE: one of SACHET, BOTTLE, TUB, CAN, BOX, GLASS JAR, CARTON, POUCH, TUBE, PACKET, BAG
- COUNTRY: country of manufacture as printed on package e.g. "MALAYSIA", "THAILAND"
- VARIANT: product variant if present e.g. "ORIGINAL", "LOW FAT", "REDUCED SALT", "PREMIUM"
- TYPE: short product category e.g. "MARGARINE", "MAYONNAISE", "SEASONING", "DETERGENT"
- FRAGRANCE_FLAVOR: scent or flavor descriptor if present e.g. "RICH", "ORIGINAL", "LEMON", "VANILLA"
- PROMOTION: any on-pack promotion text e.g. "BUY 2 FREE 1", "20% EXTRA FREE", "NEW LOWER PRICE"
- ADDONS: additional included items or features e.g. "SPOON INCLUDED", "FREE RECIPE BOOK"
- TAGLINE: short promotional tagline or slogan e.g. "TASTE THE DIFFERENCE", "LOVED BY FAMILIES"

Rules:
- UPPERCASE for all string values
- Use "" for any field not visible or not applicable
- NEVER use null
- BARCODE must be digits only
- WEIGHT must have no space between number and unit"""

TRIGGER = 'Extract the product data from this image and return only the JSON object. Use "" for missing fields. Never use null.'

# ─────────────────────────────────────────────
# SUPABASE HELPERS
# ─────────────────────────────────────────────
def get_supabase_client():
    url = st.session_state.get("supabase_url", SUPABASE_URL)
    key = st.session_state.get("supabase_key", SUPABASE_KEY)
    if url == SUPABASE_URL and key == SUPABASE_KEY:
        return supabase_client
    if url and key:
        try:
            return create_client(url, key)
        except Exception:
            return None
    return None


def supabase_upsert_product(imdb_row):
    """
    Upsert a product into the Supabase imdb_products table.
    Before inserting check if exists by barcode (if not empty) or by brand+weight+packaging_type.
    If exists increment scan_count and set is_duplicate=true and merge empty fields.
    If not exists insert as new record.
    """
    client = get_supabase_client()
    if not client:
        return False, "Supabase credentials not configured."
    
    barcode = imdb_row.get("BARCODE", "").strip()
    item_name = imdb_row.get("ITEM NAME", "").strip()
    brand = imdb_row.get("BRAND", "").strip()
    weight = imdb_row.get("WEIGHT", "").strip()
    pkg = imdb_row.get("PACKAGING TYPE", "").strip()
    
    db_fields = {
        "item_name":        item_name,
        "barcode":          barcode,
        "manufacturer":     imdb_row.get("MANUFACTURER", "").strip(),
        "brand":            brand,
        "weight":           weight,
        "packaging_type":   pkg,
        "country":          imdb_row.get("COUNTRY", "").strip(),
        "variant":          imdb_row.get("VARIANT", "").strip(),
        "type":             imdb_row.get("TYPE", "").strip(),
        "fragrance_flavor": imdb_row.get("FRAGRANCE FLAVOR", "").strip(),
        "promotion":        imdb_row.get("PROMOTION", "").strip(),
        "addons":           imdb_row.get("ADDONS", "").strip(),
        "tagline":          imdb_row.get("TAGLINE", "").strip(),
    }
    
    existing_records = None
    try:
        if barcode:
            resp = client.table("imdb_products").select("*").eq("barcode", barcode).execute()
            existing_records = resp.data
        else:
            if brand or weight or pkg:
                resp = client.table("imdb_products").select("*").eq("brand", brand).eq("weight", weight).eq("packaging_type", pkg).execute()
                existing_records = resp.data
                
        if existing_records and len(existing_records) > 0:
            existing = existing_records[0]
            scan_count = int(existing.get("scan_count") or 1) + 1
            
            # Merge empty fields: if existing value is empty/null, use new value
            updated_fields = {}
            for key, val in db_fields.items():
                existing_val = existing.get(key)
                if existing_val is None or str(existing_val).strip() == "":
                    updated_fields[key] = val
                else:
                    updated_fields[key] = existing_val
            
            updated_fields["scan_count"] = scan_count
            updated_fields["is_duplicate"] = True
            import datetime
            updated_fields["last_updated"] = datetime.datetime.utcnow().isoformat()
            
            client.table("imdb_products").update(updated_fields).eq("id", existing["id"]).execute()
            return True, f"Updated (scan #{scan_count})"
        else:
            new_record = db_fields.copy()
            new_record["scan_count"] = 1
            new_record["is_duplicate"] = False
            import datetime
            new_record["last_updated"] = datetime.datetime.utcnow().isoformat()
            
            client.table("imdb_products").insert(new_record).execute()
            return True, "Inserted as new record"
    except Exception as e:
        return False, str(e)


def supabase_get_products(search_query=""):
    """
    Fetch all products from Supabase imdb_products, optionally filtered by brand or item_name.
    """
    client = get_supabase_client()
    if not client:
        return [], "Supabase credentials not configured."
    try:
        query = client.table("imdb_products").select("*").order("last_updated", desc=True)
        if search_query:
            query = query.or_(f"brand.ilike.*{search_query}*,item_name.ilike.*{search_query}*")
        resp = query.execute()
        return resp.data, None
    except Exception as e:
        return [], str(e)


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
class NamedBytesIO(io.BytesIO):
    """File-like wrapper for captured image bytes with a custom name and size."""
    def __init__(self, content, name):
        super().__init__(content)
        self.name = name
        self.size = len(content)


def preprocess_image(image_file, max_size=(1024, 1024)):
    """Resize and base64-encode an image. Use smaller max_size for local models."""
    if hasattr(image_file, "seek"):
        image_file.seek(0)
    img = Image.open(image_file)
    if img.mode != "RGB":
        img = img.convert("RGB")
    img.thumbnail(max_size, Image.LANCZOS)
    buffer = io.BytesIO()
    img.save(buffer, format="JPEG", quality=85)
    buffer.seek(0)
    return base64.standard_b64encode(buffer.read()).decode("utf-8")


GROQ_API_KEY       = ""
OPENROUTER_API_KEY = ""
GROQ_DELAY_SEC     = 6                       # Minimum seconds between Groq/OpenRouter requests


def extract_via_groq(b64_image, api_key, max_retries=5):
    """
    Send image to Groq chat-completions endpoint (meta-llama/llama-4-scout-17b-16e-instruct).
    Exponential backoff on 429 errors: 30s, 60s, 120s, 240s, 480s.
    """
    url = "https://api.groq.com/openai/v1/chat/completions"
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    payload = {
        "model": "meta-llama/llama-4-scout-17b-16e-instruct",
        "max_tokens": 1000,
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": [
                {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64_image}"}},
                {"type": "text", "text": TRIGGER}
            ]}
        ]
    }
    for attempt in range(max_retries):
        resp = requests.post(url, json=payload, headers=headers, timeout=90)
        if resp.status_code == 429:
            wait = 30 * (2 ** attempt)
            st.warning(f"⏳ Groq rate limit hit. Waiting {wait}s before retry (attempt {attempt + 1}/{max_retries})...")
            time.sleep(wait)
            continue
        resp.raise_for_status()
        content = resp.json()["choices"][0]["message"]["content"].strip()
        content = re.sub(r"^```json\s*", "", content)
        content = re.sub(r"\s*```$", "", content)
        return json.loads(content)
    raise Exception(f"Groq API returned 429 after {max_retries} retries.")


def extract_via_openrouter(b64_image, api_key, max_retries=5):
    """
    Send image to OpenRouter chat-completions endpoint (google/gemma-4-26b-a4b-it:free).
    Exponential backoff on 429 errors: 30s, 60s, 120s, 240s, 480s.
    """
    url = "https://openrouter.ai/api/v1/chat/completions"
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    payload = {
        "model": "google/gemma-4-26b-a4b-it:free",
        "max_tokens": 1000,
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": [
                {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64_image}"}},
                {"type": "text", "text": TRIGGER}
            ]}
        ]
    }
    for attempt in range(max_retries):
        resp = requests.post(url, json=payload, headers=headers, timeout=90)
        if resp.status_code == 429:
            wait = 30 * (2 ** attempt)
            st.warning(f"⏳ OpenRouter rate limit hit. Waiting {wait}s before retry (attempt {attempt + 1}/{max_retries})...")
            time.sleep(wait)
            continue
        resp.raise_for_status()
        content = resp.json()["choices"][0]["message"]["content"].strip()
        content = re.sub(r"^```json\s*", "", content)
        content = re.sub(r"\s*```$", "", content)
        return json.loads(content)
    raise Exception(f"OpenRouter API returned 429 after {max_retries} retries.")


def aggregate(records):
    merged = {}
    fields = [
        "ITEM_NAME", "BARCODE", "MANUFACTURER", "BRAND", "WEIGHT",
        "PACKAGING_TYPE", "COUNTRY", "VARIANT", "TYPE",
        "FRAGRANCE_FLAVOR", "PROMOTION", "ADDONS", "TAGLINE",
    ]
    for col in fields:
        values = [r.get(col, "").strip() for r in records if r.get(col, "").strip()]
        if not values:
            merged[col] = ""
        else:
            count = Counter(values)
            max_count = count.most_common(1)[0][1]
            candidates = [v for v, c in count.items() if c == max_count]
            merged[col] = max(candidates, key=len)
    return merged


def validate(record):
    record["BARCODE"] = re.sub(r"[^\d]", "", record.get("BARCODE", ""))
    weight = record.get("WEIGHT", "").upper()
    weight = re.sub(r"(\d)\s+(G|KG|ML|L|MG)\b", r"\1\2", weight)
    record["WEIGHT"] = weight
    for field in ["ITEM_NAME", "MANUFACTURER", "BRAND", "PACKAGING_TYPE",
                  "COUNTRY", "VARIANT", "TYPE", "FRAGRANCE_FLAVOR",
                  "PROMOTION", "ADDONS", "TAGLINE"]:
        record[field] = record.get(field, "").upper().strip()
    return record


def to_imdb(record):
    return {
        "ITEM NAME":        record.get("ITEM_NAME", ""),
        "BARCODE":          record.get("BARCODE", ""),
        "MANUFACTURER":     record.get("MANUFACTURER", ""),
        "BRAND":            record.get("BRAND", ""),
        "WEIGHT":           record.get("WEIGHT", ""),
        "PACKAGING TYPE":   record.get("PACKAGING_TYPE", ""),
        "COUNTRY":          record.get("COUNTRY", ""),
        "VARIANT":          record.get("VARIANT", ""),
        "TYPE":             record.get("TYPE", ""),
        "FRAGRANCE FLAVOR": record.get("FRAGRANCE_FLAVOR", ""),
        "PROMOTION":        record.get("PROMOTION", ""),
        "ADDONS":           record.get("ADDONS", ""),
        "TAGLINE":          record.get("TAGLINE", ""),
    }


def build_excel(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IMDB"
    thin  = Side(style='thin', color='000000')
    thick = Side(style='medium', color='000000')
    ws.append(IMDB_COLS)
    for col_idx in range(1, len(IMDB_COLS)+1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(name='Calibri', bold=True, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=thick, right=thick, top=thick, bottom=thick)
    ws.row_dimensions[1].height = 20
    for row_idx, row in enumerate(rows, 2):
        for col_idx, col in enumerate(IMDB_COLS, 1):
            val = row.get(col, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val else None)
            cell.font = Font(name='Calibri', size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = Border(left=Side(style='thin', color='000000'),
                                 right=Side(style='thin', color='000000'),
                                 top=Side(style='thin', color='000000'),
                                 bottom=Side(style='thin', color='000000'))
        ws.row_dimensions[row_idx].height = 15
    col_widths = [18, 14, 24, 32, 16, 55, 14, 16, 18, 50]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ─────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────
if "results" not in st.session_state:
    st.session_state.results = []
if "edited_df" not in st.session_state:
    st.session_state.edited_df = None
if "groq_api_key" not in st.session_state:
    st.session_state.groq_api_key = GROQ_API_KEY
if "openrouter_api_key" not in st.session_state:
    st.session_state.openrouter_api_key = OPENROUTER_API_KEY
if "engine" not in st.session_state:
    st.session_state.engine = "Groq API"
if "camera_snapshots" not in st.session_state:
    st.session_state.camera_snapshots = []
if "last_added_photo" not in st.session_state:
    st.session_state.last_added_photo = None
if "supabase_url" not in st.session_state:
    st.session_state.supabase_url = os.environ.get("SUPABASE_URL", "")
if "supabase_key" not in st.session_state:
    st.session_state.supabase_key = os.environ.get("SUPABASE_KEY", "")


# ─────────────────────────────────────────────
# ENGINE CONFIG (MOVED FROM SIDEBAR)
# ─────────────────────────────────────────────
with st.expander("⚙️ Model Configuration", expanded=False):
    col1, col2 = st.columns(2)
    with col1:
        engine = st.selectbox(
            "Choose model",
            ["Groq API", "OpenRouter API"],
            key="engine"
        )
    with col2:
        if engine == "Groq API":
            st.text_input(
                "Groq API Key",
                type="password",
                placeholder="gsk_...",
                key="groq_api_key"
            )
        elif engine == "OpenRouter API":
            st.text_input(
                "OpenRouter API Key",
                type="password",
                placeholder="sk-or-v1-...",
                key="openrouter_api_key"
            )
    st.markdown("---")
    col3, col4 = st.columns(2)
    with col3:
        st.text_input(
            "Supabase URL",
            placeholder="https://xxxx.supabase.co",
            key="supabase_url"
        )
    with col4:
        st.text_input(
            "Supabase API Key",
            type="password",
            placeholder="your-supabase-anon-key",
            key="supabase_key"
        )

# Resolve the active api_key from the correct session state slot
if st.session_state.engine == "Groq API":
    api_key = st.session_state.groq_api_key
else:
    api_key = st.session_state.openrouter_api_key

# ─────────────────────────────────────────────
# VIEW ROUTING (Pipeline vs Item Master Database)
# ─────────────────────────────────────────────
current_view = st.query_params.get("view", "pipeline")

# ─────────────────────────────────────────────
# STATE DETECTION FOR PIPELINE TABS
# ─────────────────────────────────────────────
active_step = 1
uploaded_files_in_state = st.session_state.get("uploader", [])
camera_snapshots_in_state = st.session_state.get("camera_snapshots", [])
if uploaded_files_in_state or camera_snapshots_in_state:
    active_step = 2
    if st.session_state.get("results"):
        active_step = 5

# ─────────────────────────────────────────────
# HEADER (NAVBAR)
# ─────────────────────────────────────────────
navbar_placeholder = st.empty()
with navbar_placeholder.container():
    render_navbar(active_step, engine, current_view)

# ─────────────────────────────────────────────
# ═══════════════════════════════════════════════
# VIEW: ITEM MASTER DATABASE
# ═══════════════════════════════════════════════
if current_view == "database":
    st.markdown("<h2 style='font-family: \"Playfair Display\", Georgia, serif; font-weight: 800; font-size: 2.2rem; color: #0f172a; margin-top: 1rem; margin-bottom: 0.25rem;'>📊 Item Master Database</h2>", unsafe_allow_html=True)
    st.markdown("<p style='font-family: \"Outfit\", sans-serif; color: #64748b; font-size: 1.05rem; margin-bottom: 1.5rem;'>Browse, search, and manage all previously extracted products stored in Supabase.</p>", unsafe_allow_html=True)

    client = get_supabase_client()
    if not client:
        st.warning("⚙️ Please configure your **Supabase URL** and **API Key** in the **⚙️ Model Configuration** expander above or in Streamlit secrets to enable database features.")
    else:
        # ── Search bar ──────────────────────────────────────────────────
        col_search, col_refresh = st.columns([3, 1])
        with col_search:
            search_query = st.text_input(
                "🔍 Search by Brand or Product Name",
                placeholder="e.g. KNORR or CHICKEN STOCK...",
                label_visibility="collapsed"
            )
        with col_refresh:
            do_refresh = st.button("🔄 Refresh", use_container_width=True)

        # ── Fetch data ──────────────────────────────────────────────────
        rows, fetch_err = supabase_get_products(search_query)

        if fetch_err:
            st.error(f"❌ Could not load data from Supabase: {fetch_err}")
        elif not rows:
            st.info("📭 No products found. Run the pipeline and sync results to populate the database.")
            st.markdown("---")
            if st.button("🗑️ Clear Database (Force delete all)", type="secondary", use_container_width=True):
                try:
                    client.table("imdb_products").delete().neq("item_name", "NON_EXISTENT_VAL_THAT_WILL_NEVER_EXIST").execute()
                    st.success("✅ Database cleared successfully!")
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Error clearing database: {e}")
        else:
            # ── Summary chips ───────────────────────────────────────────
            total_products = len(rows)
            total_scans = sum(r.get("scan_count", 1) for r in rows)
            total_duplicates = sum(1 for r in rows if r.get("is_duplicate"))
            st.markdown(
                f'<div style="margin-bottom:1rem;">'
                f'<span class="db-stat-chip">📦 {total_products} Products</span>'
                f'<span class="db-stat-chip">📸 {total_scans} Total Scans</span>'
                f'<span class="db-stat-chip">⚠️ {total_duplicates} Duplicate Products</span>'
                f'</div>',
                unsafe_allow_html=True
            )

            # ── Build display DataFrame ─────────────────────────────────
            db_rows = []
            for r in rows:
                bc = r.get("barcode", "")
                is_dup = bool(r.get("is_duplicate", False))
                last_upd = r.get("last_updated", "")
                if last_upd and "T" in last_upd:
                    last_upd = last_upd.replace("T", " ")[:16]
                db_rows.append({
                    "Product Name":    r.get("item_name", ""),
                    "Brand":           r.get("brand", ""),
                    "Barcode":         bc,
                    "Weight":          r.get("weight", ""),
                    "Type":            r.get("type", ""),
                    "Country":         r.get("country", ""),
                    "Packaging":       r.get("packaging_type", ""),
                    "Scan Count":      r.get("scan_count", 1),
                    "is_duplicate":    "Yes" if is_dup else "No",
                    "Last Updated":    last_upd,
                })

            db_df = pd.DataFrame(db_rows)
            
            # Highlight duplicate rows in red (light red background and dark red text)
            def highlight_rows(row):
                is_duplicate = row["is_duplicate"] == "Yes"
                return ['background-color: #fee2e2; color: #991b1b; border: 1px solid #fca5a5;' if is_duplicate else '' for _ in row]
            
            styled_df = db_df.style.apply(highlight_rows, axis=1)
            
            st.dataframe(
                styled_df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Scan Count":   st.column_config.NumberColumn("Scan Count", format="%d 📸"),
                    "is_duplicate": st.column_config.TextColumn("is_duplicate"),
                    "Last Updated": st.column_config.TextColumn("Last Updated"),
                }
            )

            st.markdown("---")
            if st.button("🗑️ Clear Database", type="secondary", use_container_width=True):
                try:
                    client.table("imdb_products").delete().neq("item_name", "NON_EXISTENT_VAL_THAT_WILL_NEVER_EXIST").execute()
                    st.success("✅ Database cleared successfully!")
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Error clearing database: {e}")

    st.stop()   # Do not render pipeline below

# ═══════════════════════════════════════════════
# VIEW: PIPELINE (default)
# ═══════════════════════════════════════════════
st.markdown("<h2 style='font-family: \"Playfair Display\", Georgia, serif; font-weight: 800; font-size: 2.2rem; color: #0f172a; margin-top: 1rem; margin-bottom: 0.25rem;'>🏷️ IMDB Auto-Fill</h2>", unsafe_allow_html=True)
st.markdown("<p style='font-family: \"Outfit\", sans-serif; color: #64748b; font-size: 1.05rem; margin-bottom: 1.5rem;'>Welcome back! Streamline retail master data cataloging with AI-driven extraction.</p>", unsafe_allow_html=True)

# Render the page-level progress stepper
stepper_placeholder = st.empty()
with stepper_placeholder.container():
    render_pipeline_stepper(active_step)

st.divider()

# ─────────────────────────────────────────────
# STEP 1 — ACQUIRE PRODUCT IMAGES
# ─────────────────────────────────────────────
st.markdown("### Step 1 — Acquire Product Images")
st.caption("Provide product images by uploading files or capturing snapshots using your device camera.")

tab_upload, tab_camera = st.tabs(["📁 Upload Files", "📸 Capture Snapshots"])

with tab_upload:
    uploaded_files = st.file_uploader(
        "Drop images here or click to browse",
        type=["jpg", "jpeg", "png"],
        accept_multiple_files=True,
        key="uploader"
    )

with tab_camera:
    col_pid, col_clear = st.columns([2, 1])
    with col_pid:
        camera_product_id = st.text_input(
            "Product ID / Prefix",
            value="PRODUCT1",
            help="Specify a prefix (e.g. S12345) to group your camera snapshots."
        ).strip().upper()
    with col_clear:
        st.markdown("<div style='height: 28px;'></div>", unsafe_allow_html=True)
        if st.button("🗑️ Clear Snapshots", use_container_width=True):
            st.session_state.camera_snapshots = []
            st.success("Snapshot queue cleared!")
            st.rerun()

    camera_photo = st.camera_input("Take a snapshot of the product or product parts", key="camera_widget")

    if camera_photo is None:
        st.session_state.last_added_photo = None

    if camera_photo is not None:
        if not camera_product_id:
            st.warning("⚠️ Please specify a Product ID / Prefix before saving.")
        else:
            photo_bytes = camera_photo.getvalue()
            # Check if this photo was already added to prevent repeat clicks
            is_already_added = (st.session_state.last_added_photo is not None and 
                                st.session_state.last_added_photo == photo_bytes)
            
            if is_already_added:
                st.success("✅ Snapshot successfully added to queue! Tap 'Clear photo' below to capture the next angle/part.")
            else:
                if st.button(f"➕ Add Snapshot to {camera_product_id} Queue", use_container_width=True):
                    timestamp = int(time.time())
                    filename = f"{camera_product_id}_snap_{timestamp}.jpg"
                    
                    # Wrap photo bytes in NamedBytesIO
                    snap_file = NamedBytesIO(photo_bytes, filename)
                    st.session_state.camera_snapshots.append(snap_file)
                    st.session_state.last_added_photo = photo_bytes
                    st.toast(f"✅ Saved `{filename}`!")
                    st.rerun()

    # Display camera snapshot queue grid if any exist
    if st.session_state.camera_snapshots:
        st.markdown("<h4 style='margin-top: 1.5rem;'>📸 Captured Snapshots Queue</h4>", unsafe_allow_html=True)
        snap_cols = st.columns(2)
        for idx, snap in enumerate(st.session_state.camera_snapshots):
            with snap_cols[idx % 2]:
                with st.container(border=True):
                    st.image(snap.getvalue(), use_container_width=True)
                    st.markdown(f"<p class='snap-card-title' style='text-align: center;'>{snap.name}</p>", unsafe_allow_html=True)
                    if st.button("🗑️ Remove", key=f"del_snap_{idx}", use_container_width=True):
                        st.session_state.camera_snapshots.pop(idx)
                        st.rerun()

# Merge all files from both uploader and camera
all_files = []
if uploaded_files:
    all_files.extend(uploaded_files)
if st.session_state.camera_snapshots:
    all_files.extend(st.session_state.camera_snapshots)

if all_files:
    # Group by product ID (filename prefix)
    groups = defaultdict(list)
    for f in all_files:
        product_id = f.name.split("_")[0]
        groups[product_id].append(f)

    col1, col2, col3 = st.columns(3)
    col1.metric("Total images", len(all_files))
    col2.metric("Products detected", len(groups))
    col3.metric("Avg images/product", f"{len(all_files)/len(groups):.1f}" if groups else "0.0")

    # Show image previews
    with st.expander("Preview product images", expanded=False):
        cols = st.columns(5)
        for i, f in enumerate(all_files[:10]):
            with cols[i % 5]:
                if hasattr(f, "seek"):
                    f.seek(0)
                if hasattr(f, "getvalue"):
                    st.image(f.getvalue(), caption=f.name, use_container_width=True)
                else:
                    st.image(f, caption=f.name, use_container_width=True)

    st.divider()

    # ─────────────────────────────────────────────
    # STEP 2 — EXTRACT
    # ─────────────────────────────────────────────
    st.markdown("### Step 2 — Extract & Process")

    checkpoint_path = Path("./checkpoint.json")
    checkpoint_records = []
    done_ids = set()
    if checkpoint_path.exists():
        try:
            with open(checkpoint_path, "r", encoding="utf-8") as f:
                checkpoint_records = json.load(f)
            done_ids = {r["_product_id"] for r in checkpoint_records if "_product_id" in r}
            if done_ids:
                col_info, col_clear = st.columns([3, 1])
                with col_info:
                    st.info(f"ℹ️ **Found checkpoint on disk:** {len(done_ids)} product(s) already processed.")
                with col_clear:
                    if st.button("🗑️ Clear Checkpoint", use_container_width=True):
                        checkpoint_path.unlink(missing_ok=True)
                        st.rerun()
        except Exception as e:
            st.warning(f"⚠️ Could not read checkpoint file: {e}")

    if st.button("🚀 Run Pipeline", use_container_width=True):
        # Create a dictionary to hold results by product ID to preserve the original order of products
        results_by_product = {}
        for product_id in groups:
            if product_id in done_ids:
                matched_record = next(r for r in checkpoint_records if r.get("_product_id") == product_id)
                results_by_product[product_id] = to_imdb(matched_record)

        # Build a single flat queue of images to process across all non-checkpointed products
        queue = []
        for product_id, files in groups.items():
            if product_id in done_ids:
                continue
            for f in files:
                queue.append((product_id, f))

        total_images = len(queue)

        if total_images > 0:
            progress = st.progress(0)
            status = st.empty()
            
            extracted_by_product = defaultdict(list)
            processed_count_by_product = defaultdict(int)

            for idx, (product_id, f) in enumerate(queue):
                # Update navbar and stepper to "Extract" (active_step = 2)
                with navbar_placeholder.container():
                    render_navbar(2, engine)
                with stepper_placeholder.container():
                    render_pipeline_stepper(2)
                
                status.markdown(f"**Processing Image {idx + 1}/{total_images}:** `{f.name}` (Product `{product_id}`)...")
                
                timer_slot = st.empty()
                start = time.time()
                try:
                    b64 = preprocess_image(f, max_size=(1024, 1024))
                    # Read keys directly from session_state (persists across reruns)
                    _engine = st.session_state.get("engine", engine)
                    if _engine == "Groq API":
                        _api_key = st.session_state.get("groq_api_key", "")
                        if not _api_key:
                            raise ValueError("Groq API key is missing. Please enter it in Model Configuration.")
                        record = extract_via_groq(b64, _api_key)
                    elif _engine == "OpenRouter API":
                        _api_key = st.session_state.get("openrouter_api_key", "")
                        if not _api_key:
                            raise ValueError("OpenRouter API key is missing. Please enter it in Model Configuration.")
                        record = extract_via_openrouter(b64, _api_key)
                    else:
                        raise ValueError("No valid engine selected.")

                    elapsed = time.time() - start
                    timer_slot.success(f"✅ `{f.name}` done in {elapsed:.1f}s")
                    extracted_by_product[product_id].append(record)
                except Exception as e:
                    timer_slot.warning(f"⚠️ Could not extract from `{f.name}`: {e}")

                # Enforce minimum gap between API calls — only sleep what's left of the window
                if idx < total_images - 1:
                    elapsed_total = time.time() - start
                    remaining = GROQ_DELAY_SEC - elapsed_total
                    if remaining > 0:
                        status.markdown(f"⏳ Pacing... waiting {remaining:.1f}s before next call")
                        time.sleep(remaining)

                processed_count_by_product[product_id] += 1
                if processed_count_by_product[product_id] == len(groups[product_id]):
                    # All images for this product have been processed, let's aggregate and validate!
                    if extracted_by_product[product_id]:
                        with navbar_placeholder.container():
                            render_navbar(3, engine)
                        with stepper_placeholder.container():
                            render_pipeline_stepper(3)
                        merged = aggregate(extracted_by_product[product_id])
                        
                        with navbar_placeholder.container():
                            render_navbar(4, engine)
                        with stepper_placeholder.container():
                            render_pipeline_stepper(4)
                        merged = validate(merged)
                        
                        merged["_product_id"] = product_id
                        checkpoint_records.append(merged)
                        try:
                            with open(checkpoint_path, "w", encoding="utf-8") as f_out:
                                json.dump(checkpoint_records, f_out, ensure_ascii=False, indent=2)
                        except Exception as e:
                            st.warning(f"⚠️ Could not save checkpoint: {e}")
                            
                        results_by_product[product_id] = to_imdb(merged)

                progress.progress((idx + 1) / total_images)

            status.empty()
            progress.empty()

        # Reassemble the final results in the exact original order of the products
        results = [results_by_product[pid] for pid in groups if pid in results_by_product]
        
        st.session_state.results = results
        st.session_state.edited_df = pd.DataFrame(results, columns=IMDB_COLS)
        
        # Clean up checkpoint after successful full completion
        if checkpoint_path.exists():
            try:
                checkpoint_path.unlink()
            except Exception as e:
                st.warning(f"⚠️ Could not delete checkpoint file: {e}")

        st.markdown('<div class="success-banner">✅ Pipeline complete!</div>', unsafe_allow_html=True)

        # ── Auto-sync to Supabase ─────────────────────────────────────
        client = get_supabase_client()
        if client:
            sync_status = st.empty()
            sync_status.info("⏳ Syncing results to Supabase...")
            sync_errors = []
            for imdb_row in results:
                ok, msg = supabase_upsert_product(imdb_row)
                if not ok:
                    sync_errors.append(msg)
            if sync_errors:
                sync_status.warning(f"⚠️ Supabase sync had {len(sync_errors)} error(s): {sync_errors[0]}")
            else:
                sync_status.success(f"☁️ {len(results)} product(s) synced to Supabase!")

        # Update navbar and stepper to "Export/Preview" (active_step = 5)
        with navbar_placeholder.container():
            render_navbar(5, engine, current_view)
        with stepper_placeholder.container():
            render_pipeline_stepper(5)

# ─────────────────────────────────────────────
# STEP 3 — PREVIEW & EDIT
# ─────────────────────────────────────────────
if st.session_state.results:
    st.divider()
    st.markdown("### Step 3 — Preview & Edit")
    st.caption("Click any cell to edit before exporting.")

    edited = st.data_editor(
        st.session_state.edited_df,
        use_container_width=True,
        num_rows="dynamic",
        hide_index=True
    )
    st.session_state.edited_df = edited

    # Stats
    df = edited
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Total products", len(df))
    col2.metric("With barcode", int(df["BARCODE"].astype(bool).sum()))
    col3.metric("With manufacturer", int(df["MANUFACTURER"].astype(bool).sum()))
    col4.metric("With country", int(df["COUNTRY"].astype(bool).sum()))

    st.divider()

    # ─────────────────────────────────────────────
    # STEP 4 — EXPORT
    # ─────────────────────────────────────────────
    st.markdown("### Step 4 — Export")
    col1, col2 = st.columns(2)

    with col1:
        excel_buffer = build_excel(edited.to_dict("records"))
        st.download_button(
            label="📥 Download as Excel (.xlsx)",
            data=excel_buffer,
            file_name="IMDB_predictions.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    with col2:
        csv = edited.to_csv(index=False)
        st.download_button(
            label="📥 Download as CSV (.csv)",
            data=csv,
            file_name="IMDB_predictions.csv",
            mime="text/csv",
            use_container_width=True
        )

    # ── Manual Supabase Sync ─────────────────────────────────────────
    client = get_supabase_client()
    if client:
        st.markdown("---")
        if st.button("☁️ Sync Edits to Supabase", use_container_width=True):
            rows_to_sync = edited.to_dict("records")
            sync_errors = []
            with st.spinner(f"Syncing {len(rows_to_sync)} product(s) to Supabase..."):
                for row in rows_to_sync:
                    ok, msg = supabase_upsert_product(row)
                    if not ok:
                        sync_errors.append(msg)
            if sync_errors:
                st.warning(f"⚠️ {len(sync_errors)} error(s) during sync: {sync_errors[0]}")
            else:
                st.success(f"✅ {len(rows_to_sync)} product(s) synced to Supabase successfully!")
    else:
        st.caption("💡 Configure Supabase credentials in **⚙️ Model Configuration** to enable cloud sync.")