"""
GDSS-Maverick Hackathon
AI-Driven Image-to-IMDB Pipeline
Stage 1: Image Ingestion & Preprocessing
Stage 2: Attribute Extraction (Groq API or OpenRouter API)
Stage 3: Aggregation & Conflict Resolution
Stage 4: Validation, Normalization & Export
"""

import os
import re
import json
import base64
import time
import requests
from pathlib import Path
from PIL import Image
from collections import defaultdict, Counter
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

# ─────────────────────────────────────────────
# CONFIGURATION — swap extraction engine here
# ─────────────────────────────────────────────
EXTRACTION_ENGINE = "groq"    # Options: "groq", "openrouter"
GROQ_API_KEY  = os.environ.get("GROQ_API_KEY", "")  # Groq API key
GROQ_DELAY_SEC = 5            # Minimum seconds between Groq requests
OPENROUTER_API_KEY = os.environ.get("OPENROUTER_API_KEY", "")  # OpenRouter API key
IMAGE_FOLDER = "./images"     # Folder containing your product images
OUTPUT_FILE  = "./IMDB_predictions.xlsx"
MAX_IMAGE_SIZE = (1024, 1024) # Resize images larger than this before sending

# 13 IMDB columns
IMDB_COLS = [
    "ITEM NAME",
    "BARCODE",
    "MANUFACTURER",
    "BRAND",
    "WEIGHT",
    "PACKAGING TYPE",
    "COUNTRY",
    "VARIANT",
    "TYPE",
    "FRAGRANCE FLAVOR",
    "PROMOTION",
    "ADDONS",
    "TAGLINE",
]

# ─────────────────────────────────────────────
# STAGE 1: IMAGE INGESTION & PREPROCESSING
# ─────────────────────────────────────────────
def load_and_group_images(folder):
    """
    Load all images from folder, group by product ID (filename prefix).
    Returns dict: { product_id: [image_path, ...] }
    """
    folder = Path(folder)
    if not folder.exists():
        raise FileNotFoundError(f"Image folder not found: {folder}")

    groups = defaultdict(list)
    supported = {".jpg", ".jpeg", ".png", ".webp"}

    for img_path in sorted(folder.iterdir()):
        if img_path.suffix.lower() not in supported:
            continue
        # Product ID = everything before the first underscore + number group
        # e.g. S221234199_550719013.jpg → S221234199
        parts = img_path.stem.split("_")
        product_id = parts[0]
        groups[product_id].append(img_path)

    print(f"[Stage 1] Found {sum(len(v) for v in groups.values())} images across {len(groups)} products")
    return dict(groups)


def preprocess_image(img_path):
    """
    Resize image if too large, return as base64 string.
    """
    with Image.open(img_path) as img:
        # Convert to RGB (handles PNG with alpha channel)
        if img.mode != "RGB":
            img = img.convert("RGB")
        # Resize if larger than MAX_IMAGE_SIZE
        img.thumbnail(MAX_IMAGE_SIZE, Image.LANCZOS)
        # Save to bytes
        buffer = io.BytesIO()
        img.save(buffer, format="JPEG", quality=85)
        buffer.seek(0)
        return base64.standard_b64encode(buffer.read()).decode("utf-8")


# ─────────────────────────────────────────────
# STAGE 2: ATTRIBUTE EXTRACTION
# ─────────────────────────────────────────────
SYSTEM_PROMPT = """You are a product data extraction engine for a retail catalog system.
Your job is to extract structured product information from images of physical product packaging.

YOUR ONLY OUTPUT IS A SINGLE VALID JSON OBJECT. NO PREAMBLE. NO EXPLANATION. NO MARKDOWN. NO CODE BLOCKS. JUST THE RAW JSON.

Extract the following 13 fields:
- ITEM_NAME: full descriptive product name for catalog e.g. "KNORR CHICKEN STOCK CUBE 20G"
- BARCODE: numeric barcode digits only, no dashes or spaces
- MANUFACTURER: full company name that manufactures the product
- BRAND: brand name exactly as shown on package e.g. "KNORR", "SUNLIGHT"
- WEIGHT: net weight or volume with unit, no spaces e.g. "80G", "500ML", "1.5KG"
- PACKAGING_TYPE: one of SACHET, BOTTLE, TUB, CAN, BOX, GLASS JAR, CARTON, POUCH, TUBE, PACKET, BAG
- COUNTRY: country of manufacture as printed on package e.g. "MALAYSIA", "THAILAND"
- VARIANT: product variant if present e.g. "ORIGINAL", "LOW FAT", "REDUCED SALT", "PREMIUM"
- TYPE: short product category e.g. "MARGARINE", "MAYONNAISE", "SEASONING", "DETERGENT"
- FRAGRANCE_FLAVOR: scent or flavor descriptor if present e.g. "RICH", "ORIGINAL", "LEMON", "VANILLA"
- PROMOTION: any on-pack promotion text e.g. "BUY 2 FREE 1", "20% EXTRA FREE", "NEW LOWER PRICE"
- ADDONS: additional included items or features e.g. "SPOON INCLUDED", "FREE RECIPE BOOK"
- TAGLINE: short promotional tagline or slogan e.g. "TASTE THE DIFFERENCE", "LOVED BY FAMILIES"

Rules:
- Use UPPERCASE for all string values
- Use "" for any field not visible or not applicable
- NEVER use null
- BARCODE must be digits only
- WEIGHT must have no space between number and unit"""

TRIGGER = "Extract the product data from this image and return only the JSON object. Use \"\" for missing fields. Never use null."



def extract_via_groq(image_paths, max_retries=5):
    """
    ── GROQ API SWAP POINT ──
    Send images to Groq's OpenAI-compatible chat-completions endpoint using
    meta-llama/llama-4-scout-17b-16e-instruct with vision support.
    - Mandatory per-request delay controlled by GROQ_DELAY_SEC.
    - Exponential backoff starting at 30s on 429 rate-limit errors, max 5 retries.
    Returns a list of result dicts identical to the other extractors.
    """
    url = "https://api.groq.com/openai/v1/chat/completions"
    headers = {
        "Authorization": f"Bearer {GROQ_API_KEY}",
        "Content-Type": "application/json",
    }
    results = []
    for img_path in image_paths:
        print(f"  [Groq API] Extracting: {img_path.name}")
        b64 = preprocess_image(img_path)
        payload = {
            "model": "meta-llama/llama-4-scout-17b-16e-instruct",
            "messages": [
                {"role": "system", "content": SYSTEM_PROMPT},
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:image/jpeg;base64,{b64}"
                            },
                        },
                        {"type": "text", "text": TRIGGER},
                    ],
                },
            ],
            "max_tokens": 1000,
        }
        success = False
        for attempt in range(max_retries):
            try:
                time.sleep(3)  # Sleep 3 seconds before every API call to stay under 30 RPM limit
                resp = requests.post(url, json=payload, headers=headers, timeout=90)
                if resp.status_code == 429:
                    wait = 30 * (2 ** attempt)  # 30s, 60s, 120s, 240s, 480s
                    print(
                        f"  [Groq API] Rate limited. Waiting {wait}s before retry "
                        f"(attempt {attempt + 1}/{max_retries})..."
                    )
                    time.sleep(wait)
                    continue
                resp.raise_for_status()
                content = resp.json()["choices"][0]["message"]["content"].strip()
                # Strip markdown fences if present
                content = re.sub(r"^```json\s*", "", content)
                content = re.sub(r"\s*```$", "", content)
                parsed = json.loads(content)
                parsed["source_file"] = img_path.name
                results.append(parsed)
                success = True
                break
            except Exception as e:
                print(f"  [Groq API] ERROR on {img_path.name} (attempt {attempt + 1}): {e}")
                if attempt < max_retries - 1:
                    time.sleep(30 * (2 ** attempt))
                else:
                    results.append({"source_file": img_path.name, "error": str(e)})
        if not success and not any(r.get("source_file") == img_path.name for r in results):
            results.append(
                {"source_file": img_path.name,
                 "error": f"Failed after {max_retries} retries (429 rate limit)"}
            )
    return results


def extract_via_openrouter(image_paths, max_retries=5):
    """
    ── OPENROUTER API SWAP POINT ──
    Send images to OpenRouter's OpenAI-compatible chat-completions endpoint using
    meta-llama/llama-4-maverick:free with vision support.
    - Mandatory per-request delay controlled by GROQ_DELAY_SEC (shared pacing constant).
    - Exponential backoff starting at 30s on 429 rate-limit errors, max 5 retries.
    Returns a list of result dicts identical to the other extractors.
    """
    url = "https://openrouter.ai/api/v1/chat/completions"
    headers = {
        "Authorization": f"Bearer {OPENROUTER_API_KEY}",
        "Content-Type": "application/json",
    }
    results = []
    for img_path in image_paths:
        print(f"  [OpenRouter API] Extracting: {img_path.name}")
        b64 = preprocess_image(img_path)
        payload = {
            "model": "google/gemma-4-26b-a4b-it:free",
            "messages": [
                {"role": "system", "content": SYSTEM_PROMPT},
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:image/jpeg;base64,{b64}"
                            },
                        },
                        {"type": "text", "text": TRIGGER},
                    ],
                },
            ],
            "max_tokens": 1000,
        }
        # Pace requests to stay under rate limits
        time.sleep(GROQ_DELAY_SEC)
        success = False
        for attempt in range(max_retries):
            try:
                resp = requests.post(url, json=payload, headers=headers, timeout=90)
                if resp.status_code == 429:
                    wait = 30 * (2 ** attempt)  # 30s, 60s, 120s, 240s, 480s
                    print(
                        f"  [OpenRouter API] Rate limited. Waiting {wait}s before retry "
                        f"(attempt {attempt + 1}/{max_retries})..."
                    )
                    time.sleep(wait)
                    continue
                resp.raise_for_status()
                content = resp.json()["choices"][0]["message"]["content"].strip()
                # Strip markdown fences if present
                content = re.sub(r"^```json\s*", "", content)
                content = re.sub(r"\s*```$", "", content)
                parsed = json.loads(content)
                parsed["source_file"] = img_path.name
                results.append(parsed)
                success = True
                break
            except Exception as e:
                print(f"  [OpenRouter API] ERROR on {img_path.name} (attempt {attempt + 1}): {e}")
                if attempt < max_retries - 1:
                    time.sleep(30 * (2 ** attempt))
                else:
                    results.append({"source_file": img_path.name, "error": str(e)})
        if not success and not any(r.get("source_file") == img_path.name for r in results):
            results.append(
                {"source_file": img_path.name,
                 "error": f"Failed after {max_retries} retries (429 rate limit)"}
            )
    return results


def extract(image_paths):
    """Main extraction router — switches between Groq and OpenRouter APIs."""
    engine = EXTRACTION_ENGINE.lower().strip()
    if engine == "groq":
        return extract_via_groq(image_paths)
    elif engine == "openrouter":
        return extract_via_openrouter(image_paths)
    else:
        print(f"[Warning] Unknown engine '{EXTRACTION_ENGINE}', falling back to Groq.")
        return extract_via_groq(image_paths)


# ─────────────────────────────────────────────
# STAGE 3: AGGREGATION & CONFLICT RESOLUTION
# ─────────────────────────────────────────────
EXTRACT_COLS = [
    "ITEM_NAME", "BARCODE", "MANUFACTURER", "BRAND", "WEIGHT",
    "PACKAGING_TYPE", "COUNTRY", "VARIANT", "TYPE",
    "FRAGRANCE_FLAVOR", "PROMOTION", "ADDONS", "TAGLINE",
]

def aggregate(records):
    """
    Merge multiple image records for same product into one master record.
    Strategy: majority vote; ties resolved by longest value.
    Missing fields filled from any image that has them.
    """
    merged = {}
    for col in EXTRACT_COLS:
        values = [r.get(col, "").strip() for r in records if r.get(col, "").strip()]
        if not values:
            merged[col] = ""
        else:
            count = Counter(values)
            max_count = count.most_common(1)[0][1]
            candidates = [v for v, c in count.items() if c == max_count]
            merged[col] = max(candidates, key=len)
    return merged


# ─────────────────────────────────────────────
# STAGE 4: VALIDATION & NORMALIZATION
# ─────────────────────────────────────────────
def validate_and_normalize(record):
    """
    Clean and validate each field against expected formats.
    """
    # BARCODE: digits only
    record["BARCODE"] = re.sub(r"[^\d]", "", record.get("BARCODE", ""))

    # WEIGHT: uppercase, no space between number and unit
    weight = record.get("WEIGHT", "").upper()
    weight = re.sub(r"(\d)\s+(G|KG|ML|L|MG)\b", r"\1\2", weight)
    record["WEIGHT"] = weight

    # String fields: uppercase and strip
    for field in ["ITEM_NAME", "MANUFACTURER", "BRAND", "PACKAGING_TYPE",
                  "COUNTRY", "VARIANT", "TYPE", "FRAGRANCE_FLAVOR",
                  "PROMOTION", "ADDONS", "TAGLINE"]:
        record[field] = record.get(field, "").upper().strip()

    return record


def to_imdb_row(record):
    """Map internal field names to the 13 IMDB column names."""
    return {
        "ITEM NAME":        record.get("ITEM_NAME", ""),
        "BARCODE":          record.get("BARCODE", ""),
        "MANUFACTURER":     record.get("MANUFACTURER", ""),
        "BRAND":            record.get("BRAND", ""),
        "WEIGHT":           record.get("WEIGHT", ""),
        "PACKAGING TYPE":   record.get("PACKAGING_TYPE", ""),
        "COUNTRY":          record.get("COUNTRY", ""),
        "VARIANT":          record.get("VARIANT", ""),
        "TYPE":             record.get("TYPE", ""),
        "FRAGRANCE FLAVOR": record.get("FRAGRANCE_FLAVOR", ""),
        "PROMOTION":        record.get("PROMOTION", ""),
        "ADDONS":           record.get("ADDONS", ""),
        "TAGLINE":          record.get("TAGLINE", ""),
    }


# ─────────────────────────────────────────────
# EXPORT
# ─────────────────────────────────────────────
def export_excel(rows, output_path):
    """Export final IMDB rows to clean Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IMDB"

    thin  = Side(style='thin',   color='000000')
    thick = Side(style='medium', color='000000')
    cell_border   = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
    header_border = Border(left=thick, right=thick, top=thick, bottom=thick)

    header_font  = Font(name='Calibri', bold=True, size=11)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=False)
    data_font    = Font(name='Calibri', size=10)
    data_align   = Alignment(horizontal='left', vertical='center', wrap_text=False)

    ws.append(IMDB_COLS)
    for col_idx in range(1, len(IMDB_COLS)+1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = header_align
        cell.border = header_border
    ws.row_dimensions[1].height = 20

    for row_idx, row in enumerate(rows, 2):
        for col_idx, col in enumerate(IMDB_COLS, 1):
            val = row.get(col, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val else None)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = cell_border
        ws.row_dimensions[row_idx].height = 15

    col_widths = [18, 14, 24, 32, 16, 55, 14, 16, 18, 50]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'
    wb.save(output_path)
    print(f"[Export] Saved: {output_path} ({len(rows)} products)")


# ─────────────────────────────────────────────
# MAIN — run the full pipeline
# ─────────────────────────────────────────────
def run_pipeline():
    # ── Test-mode controls ──────────────────────────────────────────────────
    TEST_MODE          = False  # Set False to run full pipeline
    TEST_PRODUCT_LIMIT = 1      # Number of product groups to process in test mode
    # ────────────────────────────────────────────────────────────────────────

    CHECKPOINT_FILE = "./checkpoint.json"

    print("=" * 50)
    print("GDSS-Maverick IMDB Pipeline")
    if TEST_MODE:
        print(f"[TEST MODE] Processing first {TEST_PRODUCT_LIMIT} product(s) only")
    print("=" * 50)

    # ── Checkpoint: load already-processed results ───────────────────────────
    checkpoint_path = Path(CHECKPOINT_FILE)
    if checkpoint_path.exists():
        try:
            with open(checkpoint_path, "r", encoding="utf-8") as f:
                all_records = json.load(f)
            done_ids = {r["_product_id"] for r in all_records}
            print(f"[Checkpoint] Resuming — {len(done_ids)} product(s) already done: {sorted(done_ids)}")
        except Exception as e:
            print(f"[Checkpoint] WARNING: Could not read checkpoint file ({e}). Starting fresh.")
            all_records = []
            done_ids = set()
    else:
        all_records = []
        done_ids = set()
    # ────────────────────────────────────────────────────────────────────────

    # Stage 1: Load & group images
    groups = load_and_group_images(IMAGE_FOLDER)

    # Apply test-mode product limit
    if TEST_MODE:
        groups = dict(list(groups.items())[:TEST_PRODUCT_LIMIT])

    for product_id, image_paths in groups.items():
        # Skip products already saved in the checkpoint
        if product_id in done_ids:
            print(f"\n[Product] {product_id} — skipped (already in checkpoint)")
            continue

        print(f"\n[Product] {product_id} ({len(image_paths)} images)")

        # Stage 2: Extract
        extracted = extract(image_paths)

        # ── Raw extraction dump (always shown in TEST_MODE) ──────────────────
        if TEST_MODE:
            print("\n[TEST MODE] Raw extracted JSON before aggregation:")
            print("-" * 50)
            for i, record in enumerate(extracted, 1):
                print(f"  Image {i} ({record.get('source_file', '?')}):")
                print(json.dumps(record, indent=4, ensure_ascii=False))
            print("-" * 50)
        # ────────────────────────────────────────────────────────────────────

        # Stage 3: Aggregate
        valid = [r for r in extracted if "error" not in r]
        if not valid:
            print(f"  [Warning] No valid extractions for {product_id}")
            continue
        merged = aggregate(valid)

        # Stage 4: Validate & normalize
        merged = validate_and_normalize(merged)

        # Tag record with product ID for checkpoint tracking
        merged["_product_id"] = product_id
        all_records.append(merged)

        # ── Checkpoint: persist immediately after each successful product ─────
        try:
            with open(checkpoint_path, "w", encoding="utf-8") as f:
                json.dump(all_records, f, ensure_ascii=False, indent=2)
            print(f"  [Checkpoint] Saved ({len(all_records)} product(s) so far)")
        except Exception as e:
            print(f"  [Checkpoint] WARNING: Could not write checkpoint ({e})")
        # ────────────────────────────────────────────────────────────────────

    # Strip internal tracking key before export
    export_records = [{k: v for k, v in r.items() if k != "_product_id"} for r in all_records]

    # Map to IMDB columns
    imdb_rows = [to_imdb_row(r) for r in export_records]

    # Export
    export_excel(imdb_rows, OUTPUT_FILE)

    # ── Checkpoint: clean up after successful export ──────────────────────────
    if checkpoint_path.exists():
        try:
            checkpoint_path.unlink()
            print("[Checkpoint] Deleted checkpoint.json after successful export.")
        except Exception as e:
            print(f"[Checkpoint] WARNING: Could not delete checkpoint file ({e})")
    # ────────────────────────────────────────────────────────────────────────

    print("\n" + "=" * 50)
    print(f"Pipeline complete. {len(imdb_rows)} products exported.")
    if TEST_MODE:
        print("[TEST MODE] Set TEST_MODE = False to run the full pipeline.")
    print("=" * 50)


if __name__ == "__main__":
    run_pipeline()